from BTScraper_selenium import BT_scraper as scraper
from BTScraper_selenium import progress_scraper as progress_scraper
from openpyxl import Workbook
from openpyxl import load_workbook
import json
import os
import time

path = os.getcwd()

# path = r'C:\Users\daniel.lavrentiev\Dropbox\Screen Project\Scraper'
scraper('danlav', "Yakima12!@", path)

workbook = load_workbook(filename='Jobsites.xlsx')
sheet = workbook.active

sheet.unmerge_cells('A1:G1')
sheet.delete_rows(1)


max_row = sheet.max_row
max_column = sheet.max_column

for i in range (2, max_row+1):

    cell_obj = sheet.cell(row=i, column=1)
    cell_col = cell_obj.fill.start_color.index
    cell_obj.value = cell_col

workbook.save('Jobsites.xlsx')



houses = []
row_array = []


for j in range (1, max_row+1):
    row_array = []
    for i in range (1, max_column+1):
        cell_obj = sheet.cell(row=j, column=i)
        row_array.append(cell_obj.value)
    houses.append(row_array)



# json_str = json.dumps(houses, default=str, indent=4)

screen_path = 'screen/src'

with open(os.path.join(screen_path, 'houses.json'), 'w') as json_file:
    json.dump(houses, json_file, default=str)

os.remove("Jobsites.xlsx")

progress_scraper('danlav', "Yakima12!@", screen_path)